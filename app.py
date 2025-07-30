from flask import Flask, render_template, request, redirect, url_for
from dotenv import load_dotenv
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font
import os
import requests

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://requirement_db_user:1ZIK7d6WCcpBcrcu5HnRwK6AyE2cvV7U@dpg-d1viequr433s73fn32rg-a.oregon-postgres.render.com/requirement_db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

EXCEL_FOLDER = r"\\Server\d\Requirement_Excel"
os.makedirs(EXCEL_FOLDER, exist_ok=True)

load_dotenv()

TENANT_ID= os.getenv('TENANT_ID')
CLIENT_ID= os.getenv('CLIENT_ID')
CLIENT_SECRET= os.getenv('CLIENT_SECRET')

EXCEL_PATHS = {
    'Nomura_JAVA': os.path.join(EXCEL_FOLDER, "Nomura_Java.xlsx"),
    'Morgan_Stanley': os.path.join(EXCEL_FOLDER, "Morgan_Stanley.xlsx"),
    'Nomura_TECH': os.path.join(EXCEL_FOLDER, "Nomura_TECH.xlsx"),
    'Nomura_NonTECH': os.path.join(EXCEL_FOLDER, "Nomura_NonTECH.xlsx"),
    'Nomura_Senior': os.path.join(EXCEL_FOLDER, "Nomura_Senior.xlsx"),
    'British_Petrolium': os.path.join(EXCEL_FOLDER, "British_Petrolium.xlsx"),
    'Russell': os.path.join(EXCEL_FOLDER, "Russell.xlsx"),
    'MUFG': os.path.join(EXCEL_FOLDER, "MUFG.xlsx"),
    'Chevron': os.path.join(EXCEL_FOLDER, "Chevron.xlsx"),
    'Lufthansa': os.path.join(EXCEL_FOLDER, "Lufthansa.xlsx"),
    'Interactive_Brokers': os.path.join(EXCEL_FOLDER, "Interactive_Brokers.xlsx"),
}

COLUMN_ORDER = {
    'sequence': ['Status', 'date', 'req_id', 'req_Name', 'Mandatory_skills', 'Desirable_Skills',
                 'exp_range', 'sal_budget', 'Notice_Period', 'Total_Upload', 'Location', 'Client_Spoc', 'Remarks']
}

# TENANT_ID = '4b31b85e-bd6e-4e7e-afb1-cc86899fac33'
# CLIENT_ID = '75cfcf77-e81e-43d6-b01b-aefd9fa5dd37'
# CLIENT_SECRET = 'KNt8Q~.GVokYoZ56ViRxR-kfbSij3ovfqbqNEaTp'


class Requirement(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    client = db.Column(db.String(100))
    req_id = db.Column(db.String(100))
    req_Name = db.Column(db.String(200))
    date = db.Column(db.String(20))
    sal_budget = db.Column(db.String(100))
    Notice_Period = db.Column(db.String(100))
    Mandatory_skills = db.Column(db.Text)
    Desirable_Skills = db.Column(db.Text)
    exp_range = db.Column(db.String(100))
    Client_Spoc = db.Column(db.String(100))
    Total_Upload = db.Column(db.String(100))
    Remarks = db.Column(db.Text)
    Location = db.Column(db.String(100))
    Status = db.Column(db.String(100))


def get_access_token():
    url = f"https://login.microsoftonline.com/{TENANT_ID}/oauth2/v2.0/token"
    data = {
        'client_id': CLIENT_ID,
        'scope': 'https://graph.microsoft.com/.default',
        'client_secret': 'KNt8Q~.GVokYoZ56ViRxR-kfbSij3ovfqbqNEaTp',
        'grant_type': 'client_credentials'
    }
    response = requests.post(url, data=data)
    return response.json().get('access_token')


def append_to_excel_onedrive(row_data):
    access_token = get_access_token()
    if not access_token:
        return "❌ Failed to get token"

    headers = {
        'Authorization': f'Bearer {access_token}',
        'Content-Type': 'application/json'
    }

    body = {
        "values": [row_data]
    }

    url = "https://graph.microsoft.com/v1.0/users/{user_id}/drive/root:/Documents/Lufthansa.xlsx:/workbook/tables/Table1/rows/add"

    response = requests.post(url, headers=headers, json=body)
    if response.status_code == 201:
        print("✅ Row added to OneDrive Excel table successfully.")
    else:
        print("❌ Failed to add row to OneDrive Excel table.")
        print("Status Code:", response.status_code)
        print("Response Text:", response.text)


def append_to_excel(file_path, data_dict):
    headers = list(data_dict.keys())
    values = list(data_dict.values())

    try:
        if not os.path.exists(file_path):
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.append(values)
            wb.save(file_path)
        else:
            wb = load_workbook(file_path)
            ws = wb.active
            ws.append(values)
            wb.save(file_path)
    except InvalidFileException:
        print(f" Invalid Excel file format: {file_path}. Delete and recreate.")
        raise
    except Exception as e:
        print(f" Error writing to Excel: {e}")
        raise


@app.route('/')
def home():
    today_str = datetime.now().strftime('%d-%m-%Y')
    clients = list(EXCEL_PATHS.keys())
    notices = []

    for client in clients:
        latest = Requirement.query.filter_by(client=client, date=today_str).order_by(Requirement.id.desc()).first()
        if latest:
            notices.append({
                'client': latest.client,
                'req_id': latest.req_id,
                'req_Name': latest.req_Name,
                'date': latest.date,
                'Mandatory_skills': latest.Mandatory_skills,
                'Remarks': latest.Remarks
            })
    return render_template('home.html', clients=clients, notices=notices)


@app.route('/dashboard/<client_name>')
def dashboard(client_name):
    rows = Requirement.query.filter_by(client=client_name).order_by(Requirement.id.desc()).all()
    return render_template('dashboard.html', client=client_name, rows=rows)


@app.route('/add/<client_name>')
def add_requirement(client_name):
    today = datetime.now().strftime('%d-%m-%Y')
    return render_template('form.html', client=client_name, today=today)


@app.route('/submit/<client_name>', methods=['POST'])
def submit(client_name):
    entry = Requirement(
        client=client_name,
        req_id=request.form['req_id'],
        req_Name=request.form['req_Name'],
        date=request.form['date'],
        sal_budget=request.form['sal_budget'],
        Notice_Period=request.form['Notice_Period'],
        Mandatory_skills=request.form['Mandatory_skills'],
        Desirable_Skills=request.form['Desirable_Skills'],
        exp_range=request.form['exp_range'],
        Client_Spoc=request.form['Client_Spoc'],
        Total_Upload=request.form['Total_Upload'],
        Remarks=request.form['Remarks'],
        Location=request.form['Location'],
        Status=request.form['Status']
    )
    db.session.add(entry)
    db.session.commit()

    form_data = {
        'Status': entry.Status,
        'date': entry.date,
        'req_id': entry.req_id,
        'req_Name': entry.req_Name,
        'Mandatory_skills': entry.Mandatory_skills,
        'Desirable_Skills': entry.Desirable_Skills,
        'exp_range': entry.exp_range,
        'sal_budget': entry.sal_budget,
        'Notice_Period': entry.Notice_Period,
        'Total_Upload': entry.Total_Upload,
        'Location': entry.Location,
        'Client_Spoc': entry.Client_Spoc,
        'Remarks': entry.Remarks
    }

    append_to_excel(EXCEL_PATHS[client_name], form_data)
    append_to_excel_onedrive(list(form_data.values()))

    return redirect(url_for('dashboard', client_name=client_name))


@app.route('/edit/<client_name>/<int:entry_id>', methods=['GET', 'POST'])
def edit_entry(client_name, entry_id):
    entry = Requirement.query.get_or_404(entry_id)

    if request.method == 'POST':
        entry.req_id = request.form['req_id']
        entry.req_Name = request.form['req_Name']
        entry.date = request.form['date']
        entry.sal_budget = request.form['sal_budget']
        entry.Notice_Period = request.form['Notice_Period']
        entry.Mandatory_skills = request.form['Mandatory_skills']
        entry.Desirable_Skills = request.form['Desirable_Skills']
        entry.exp_range = request.form['exp_range']
        entry.Client_Spoc = request.form['Client_Spoc']
        entry.Total_Upload = request.form['Total_Upload']
        entry.Remarks = request.form['Remarks']
        entry.Location = request.form['Location']
        entry.Status = request.form['Status']

        db.session.commit()

        form_data = {
            'Status': entry.Status,
            'date': entry.date,
            'req_id': entry.req_id,
            'req_Name': entry.req_Name,
            'Mandatory_skills': entry.Mandatory_skills,
            'Desirable_Skills': entry.Desirable_Skills,
            'exp_range': entry.exp_range,
            'sal_budget': entry.sal_budget,
            'Notice_Period': entry.Notice_Period,
            'Total_Upload': entry.Total_Upload,
            'Location': entry.Location,
            'Client_Spoc': entry.Client_Spoc,
            'Remarks': entry.Remarks
        }
        append_to_excel(EXCEL_PATHS[client_name], form_data)
        return redirect(url_for('dashboard', client_name=client_name))

    return render_template('edit_entry.html', client=client_name, row=entry)


if __name__ == '__main__':
    with app.app_context():
        db.create_all()

    app.run(host='0.0.0.0', port=5000, debug=True)
